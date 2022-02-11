from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Tracking.xlsx')
ws = wb.active

frameNumber = 2


for row in range(2,11):
    frameNumber += 1
    for col in range(1,2):
        char = get_column_letter(col)
        ws[char + str(row)] = ws.append(['Frame' + str(frameNumber)])


wb.save('Tracking.xlsx')
